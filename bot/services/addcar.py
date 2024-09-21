import re
import threading

import telebot
from django.db.models import Q
from django.utils import timezone
from telebot.types import InlineKeyboardButton, InlineKeyboardMarkup

from conf.settings import ADMINS, CHANNEL_ID

from ..buttons.default import ask_phone, main_button, get_back
from ..models import TgUser, Product, Material, Report
from ..services.steps import USER_STEP
from django.db.models import Q
from bot.buttons.inline import create_confirmation_keyboard
import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO
import threading
import time


def is_number(text):
    try:
        float(text)
        return True
    except ValueError:
        return False
    
# Function to send a message and delete it after a delay
def send_and_delete_message(bot, chat_id, text, delay=2):
    msg = bot.send_message(chat_id=chat_id, text=text)
    time.sleep(delay)
    bot.delete_message(chat_id=chat_id, message_id=msg.message_id)

def add_product(message, bot):
    try:
        user = TgUser.objects.get(telegram_id=message.from_user.id)
        search_text = message.text.strip()
        bot.delete_message(message.chat.id, message.id)

        pr_list = Product.objects.filter(
            Q(name__icontains=search_text) | Q(itemcode__icontains=search_text)
        ).order_by('name')

        if pr_list.exists():
            markup = InlineKeyboardMarkup()
            row = []
            for product in pr_list:
                row.append(InlineKeyboardButton(text=product.name, callback_data=f"product_{product.id}"))
                if len(row) == 2:
                    markup.add(*row)
                    row = []
            if row:
                markup.add(*row)
            
            bot.edit_message_reply_markup(message.chat.id, user.edit_msg, reply_markup=markup)
        else:
            threading.Thread(target=send_and_delete_message, args=(bot, message.chat.id, "No products found. Please try a different search.", 5)).start()

    except Exception as e:
        print(e)
        bot.send_message(
            chat_id=message.chat.id,
            text="An error occurred. Please try again."
        )


def add_material(message, bot):
    try:
        user = TgUser.objects.get(telegram_id=message.from_user.id)
        search_text = message.text.strip()
        bot.delete_message(message.chat.id, message.id)
        search_words = search_text.split()
        
        query = Q()
        for word in search_words:
            query |= Q(name__icontains=word) | Q(itemcode__icontains=word)

        ml_list = Material.objects.filter(query)

        if ml_list.exists():
            markup = InlineKeyboardMarkup()
            row = []
            for material in ml_list:
                row.append(InlineKeyboardButton(text=material.name, callback_data=f"material_{material.id}"))
                if len(row) == 2:
                    markup.add(*row)
                    row = []
            if row:
                markup.add(*row)
            
            bot.edit_message_reply_markup(message.chat.id, user.edit_msg, reply_markup=markup)
        else:
            bot.send_message(
                chat_id=message.chat.id,
                text="No material found. Please try a different search."
            )

    except Exception as e:
        print(e)
        bot.send_message(
            chat_id=message.chat.id,
            text="An error occurred. Please try again."
        )



def add_measure(message, bot):
    measure = message.text
    if is_number(measure):
        user = TgUser.objects.get(telegram_id=message.from_user.id)
        report = Report.objects.get(is_confirmed=False, user=user)
        report.termoplast_measure = measure
        report.save()
        user.step = USER_STEP['ADD_WASTE']
        user.save()
        text=f"Report\n{report.date} - {report.default_value}\nMachine Number: {report.machine_num}\n\nProduct: {report.product.name}\nMaterial: {report.material.name}\n\nTermoplast vazni: {report.termoplast_measure}\n\nBrak vazni:",
        bot.edit_message_text(chat_id=message.from_user.id, message_id=user.edit_msg, text=text, parse_mode='html')
    else:
        threading.Thread(target=send_and_delete_message, args=(bot, message.chat.id, "son kiriting ❌", 2)).start()
    bot.delete_message(message.chat.id, message.id)
    
def add_waste(message, bot):
    measure = message.text
    if is_number(measure):
        user = TgUser.objects.get(telegram_id=message.from_user.id)
        report = Report.objects.get(is_confirmed=False, user=user)
        report.waste_measure = measure
        report.save()
        user.step = USER_STEP['ADD_DEFECT']
        user.save()
        text=f"Report\n{report.date} - {report.default_value}\nMachine Number: {report.machine_num}\n\nProduct: {report.product.name}\nMaterial: {report.material.name}\n\nTermoplast vazni: {report.termoplast_measure}\n\nBrak vazni: {report.waste_measure}\n\nAtxod vaznini kiriting:",
        bot.edit_message_text(chat_id=message.from_user.id, message_id=user.edit_msg, text=text, parse_mode='html')
    else:
        threading.Thread(target=send_and_delete_message, args=(bot, message.chat.id, "son kiriting ❌", 2)).start()
    bot.delete_message(message.chat.id, message.id)
    
    
def add_defect(message, bot):
    measure = message.text
    if is_number(measure):
        user = TgUser.objects.get(telegram_id=message.from_user.id)
        report = Report.objects.get(is_confirmed=False, user=user)
        report.defect_measure = measure
        report.save()
        user.step = USER_STEP['ADD_QUANTITY']
        user.save()
        text=f"Report\n{report.date} - {report.default_value}\nMachine Number: {report.machine_num}\n\nProduct: {report.product.name}\nMaterial: {report.material.name}\n\nTermoplast vazni: {report.termoplast_measure}\n\nBrak vazni: {report.waste_measure}\n\nAtxod vaznini: {report.defect_measure}\n\nMaxsulot sonini kiriting: ",
        bot.edit_message_text(chat_id=message.from_user.id, message_id=user.edit_msg, text=text, parse_mode='html')
    else:
        threading.Thread(target=send_and_delete_message, args=(bot, message.chat.id, "son kiriting ❌", 2)).start()
    bot.delete_message(message.chat.id, message.id)



def add_quantity(message, bot):
    measure = message.text
    if is_number(measure):
        user = TgUser.objects.get(telegram_id=message.from_user.id)
        report = Report.objects.get(is_confirmed=False, user=user)
        report.quantity = measure
        report.save()
        user.step = USER_STEP['FINISH_REPORT']
        user.save()
        markup = create_confirmation_keyboard(report.id)
        text=f"Report\n{report.date} - {report.default_value}\n\nProduct: {report.product.name}\nMaterial: {report.material.name}\n\nTermoplast vazni: {report.termoplast_measure}\n\nBrak vazni: {report.waste_measure}\n\nAtxod vaznini: {report.defect_measure}\n\n Maxsulot sonini:{report.quantity}",
        bot.edit_message_text(chat_id=message.from_user.id, message_id=user.edit_msg, text=text, parse_mode='html', reply_markup=markup)
    else:
        threading.Thread(target=send_and_delete_message, args=(bot, message.chat.id, "son kiriting ❌", 2)).start()
    bot.delete_message(message.chat.id, message.id)
    


def create_excel_report(reports, file_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reports"

    # Define the header
    headers = ['Apparat', 'Maxsulot', 'Toza', 'Brak', 'Atxod', 'Siryo', 'Soni', 'Smena', 'Umumiy vazni', 'TOZA+BRAK']
    ws.append(headers)

    # Populate the sheet with data from the reports
    for report in reports:
        row = [
            report.machine_num,
            report.product.name if report.product else '',  # Assuming Product has a name field
            report.termoplast_measure,
            report.waste_measure,
            report.defect_measure,
            report.material.name if report.material else '',  # Assuming Material has a name field
            report.quantity,
            report.default_value,
            report.termoplast_measure+ report.waste_measure + report.defect_measure,
            report.termoplast_measure + report.waste_measure
        ]
        ws.append(row)
        
        # Auto-size columns based on the length of the data in each column
    for col in ws.columns:
        max_length = 0
        column = col[0].column  # Get the column number
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)  # Add a little extra space
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width

    # Save the workbook to the specified file path
    wb.save(file_path)
    wb.close()